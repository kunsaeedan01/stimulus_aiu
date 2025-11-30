from typing import Iterable
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils.timezone import localtime
from .models import Application, Paper

def _coauthors_human(paper: Paper) -> str:
    items = []
    for co in paper.coauthors.all():
        details = []
        if co.is_aiu_employee:
            details.append("AIU")
        
        if co.email:
            details.append(co.email)
        
        if co.position:
            details.append(co.position)
        
        info_str = f" ({', '.join(details)})" if details else ""
        items.append(f"{co.full_name}{info_str}")
    
    return "\n".join(items)

def build_applications_xlsx(applications_qs: Iterable[Application]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    columns_config = [
        ("App ID", 12, False),
        ("Report Year", 12, False),
        ("Status", 15, True),
        ("Faculty", 30, True),
        ("Owner Name", 30, True),
        ("Owner Email", 30, False),
        ("Position", 20, True),
        ("Subdivision", 20, True),
        ("Phone", 15, False),
        ("Created At", 18, False),
        ("Admin Comment", 30, True),
        ("Paper ID", 36, False),
        ("Title", 40, True),         
        ("Journal/Source", 25, True),
        ("Indexation", 12, False),
        ("Quartile (WoS)", 15, False),
        ("Percentile (Scopus)", 18, False),
        ("DOI", 25, False),      
        ("Pub. Date", 12, False),
        ("Year", 8, False),
        ("Vol/Num/Pages", 20, False),
        ("Affiliation (AIU)", 15, False),
        ("Platonus", 15, False),
        ("Source URL", 30, False),  
        ("Coauthors", 35, True),     
        ("File Name", 25, False),    
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    align_top_wrap = Alignment(vertical="top", wrap_text=True)
    align_top_nowrap = Alignment(vertical="top", wrap_text=False)

    for col_idx, (col_name, width, _) in enumerate(columns_config, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A2"

    row_idx = 2
    for app in applications_qs:
        created_str = localtime(app.created_at).strftime("%Y-%m-%d %H:%M")
        
        faculty_disp = app.get_faculty_display() if hasattr(app, 'get_faculty_display') else app.faculty
        status_disp = app.get_status_display() if hasattr(app, 'get_status_display') else app.status

        base_data = [
            str(app.id).split('-')[0], 
            app.report_year,
            status_disp,
            faculty_disp or "",
            app.owner.full_name or "",
            app.owner.email or "",
            app.owner.position or "",
            app.owner.subdivision or "",
            app.owner.telephone or "",
            created_str,
            app.admin_comment or ""
        ]

        papers = list(app.papers.all())

        if not papers:
            full_row = base_data + [""] * 15 
            for c_i, val in enumerate(full_row, 1):
                cell = ws.cell(row=row_idx, column=c_i, value=val)
                should_wrap = columns_config[c_i-1][2]
                cell.alignment = align_top_wrap if should_wrap else align_top_nowrap
            row_idx += 1
            continue

        for p in papers:
            details_parts = []
            if p.volume: details_parts.append(f"Vol:{p.volume}")
            if p.number: details_parts.append(f"No:{p.number}")
            if p.pages: details_parts.append(f"pp.{p.pages}")
            details_str = ", ".join(details_parts)

            indexation_disp = p.get_indexation_display() if hasattr(p, 'get_indexation_display') else p.indexation
            
            paper_data = [
                str(p.id),
                p.title or "",
                p.journal_or_source or "",
                indexation_disp or "",
                p.quartile or "",
                p.percentile if p.percentile is not None else "",
                p.doi or "",
                p.publication_date.strftime("%d.%m.%Y") if p.publication_date else "",
                p.year if p.year else "",
                details_str,
                "Yes" if p.has_university_affiliation else "No",
                "Yes" if p.registered_in_platonus else "No",
                p.source_url or "",
                _coauthors_human(p),
                (p.file_upload.name.split("/")[-1] if p.file_upload else ""),
            ]

            full_row = base_data + paper_data

            for c_i, val in enumerate(full_row, 1):
                cell = ws.cell(row=row_idx, column=c_i, value=val)
                should_wrap = columns_config[c_i-1][2]
                cell.alignment = align_top_wrap if should_wrap else align_top_nowrap
                if c_i == 3: 
                    if app.status == 'approved':
                        cell.font = Font(color="006100", bold=True) 
                    elif app.status == 'submitted':
                        cell.font = Font(color="806000", bold=True) 
                    elif app.status == 'rejected':
                        cell.font = Font(color="9C0006", bold=True) 

            row_idx += 1

    io_buffer = BytesIO()
    wb.save(io_buffer)
    io_buffer.seek(0)
    return io_buffer.read()